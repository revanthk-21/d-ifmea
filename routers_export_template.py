"""
routers/export_template.py
POST /api/dfmea/export/template

Generates a styled AIAG-VDA-aligned DFMEA worksheet as an .xlsx file
and returns it as a streaming download.

Column layout follows the DFMEA Worksheet tab from the reference file:
  Structure Analysis | Failure Analysis | Risk Assessment
"""

import io
from fastapi import APIRouter
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

router = APIRouter()

# ── Colour palette (matches reference DFMEA worksheet) ───────────────────────

HEADER_FILL      = PatternFill("solid", fgColor="1F3864")   # dark navy
SUBHEADER_FILL   = PatternFill("solid", fgColor="2E75B6")   # mid blue
COL_STRUCT_FILL  = PatternFill("solid", fgColor="D9E1F2")   # light blue
COL_FAIL_FILL    = PatternFill("solid", fgColor="FCE4D6")   # light salmon
COL_RISK_FILL    = PatternFill("solid", fgColor="E2EFDA")   # light green
HIGH_RPN_FILL    = PatternFill("solid", fgColor="FF0000")   # red — RPN ≥ 200
MED_RPN_FILL     = PatternFill("solid", fgColor="FFCC00")   # amber — RPN 100-199

WHITE_BOLD  = Font(name="Arial", bold=True,  color="FFFFFF", size=9)
DARK_BOLD   = Font(name="Arial", bold=True,  color="1F3864", size=9)
NORMAL_FONT = Font(name="Arial", bold=False, color="000000", size=8)
SMALL_FONT  = Font(name="Arial", bold=False, color="444444", size=7)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="top",    wrap_text=True)

THIN_SIDE   = Side(border_style="thin",   color="B8B8B8")
MEDIUM_SIDE = Side(border_style="medium", color="1F3864")
THIN_BORDER   = Border(left=THIN_SIDE,   right=THIN_SIDE,   top=THIN_SIDE,   bottom=THIN_SIDE)
MEDIUM_BORDER = Border(left=MEDIUM_SIDE, right=MEDIUM_SIDE, top=MEDIUM_SIDE, bottom=MEDIUM_SIDE)


# ── Column definitions ────────────────────────────────────────────────────────
# (header, width, section, description)

COLUMNS = [
    # Structure Analysis
    ("Focus Element",       22, "structure", "The element under analysis"),
    ("Focus Function",      28, "structure", "Function of the focus element"),
    ("Lower Element",       22, "structure", "Lower-level element"),
    ("Lower Function",      28, "structure", "Function of the lower-level element"),
    ("Higher Element",      22, "structure", "Higher-level element"),
    ("Higher Function",     28, "structure", "Function of the higher-level element"),
    # Failure Analysis
    ("Failure Mode",        30, "failure",   "How the focus function fails"),
    ("Noise Factor",        28, "failure",   "Noise factor triggering the failure cause"),
    ("Failure Cause",       35, "failure",   "Root cause of the failure mode"),
    ("Failure Effect",      35, "failure",   "Effect on the higher-level function"),
    # Risk Assessment
    ("Prevention Methods",  32, "risk",      "Current design prevention actions"),
    ("Detection Methods",   32, "risk",      "Current detection / verification methods"),
    ("S",                   6,  "risk",      "Severity (1–10)"),
    ("O",                   6,  "risk",      "Occurrence (1–10)"),
    ("D",                   6,  "risk",      "Detection (1–10)"),
    ("RPN",                 8,  "risk",      "Risk Priority Number = S × O × D"),
    ("AP",                  6,  "risk",      "Action Priority (H/M/L)"),
]

SECTION_FILLS = {
    "structure": COL_STRUCT_FILL,
    "failure":   COL_FAIL_FILL,
    "risk":      COL_RISK_FILL,
}


# ── Request model ─────────────────────────────────────────────────────────────

class ExportRow(BaseModel):
    focus_element:      str
    focus_function:     str
    failure_mode:       str
    failure_effect:     str
    severity:           int | None
    failure_cause:      str
    prevention_methods: str
    detection_methods:  str
    detection:          int | None
    rpn:                int | None
    # These come from the full DFMEARow but export endpoint only needs a subset.
    # Add optional fields so the payload from the frontend doesn't break.
    lower_element:      str = ""
    lower_function:     str = ""
    higher_element:     str = ""
    higher_function:    str = ""
    noise_factor:       str = ""
    occurrence:         int | None = None
    action_priority:    str = ""


class ExportRequest(BaseModel):
    rows: list[ExportRow]


# ── Workbook builder ──────────────────────────────────────────────────────────

def _build_workbook(rows: list[ExportRow]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "DFMEA Worksheet"
    ws.sheet_view.showGridLines = False

    num_cols = len(COLUMNS)

    # ── Row 1: Title banner ────────────────────────────────────────────────
    ws.row_dimensions[1].height = 28
    title_cell = ws.cell(row=1, column=1,
                         value="DFMEA Worksheet  —  AIAG-VDA 2019 / Bosch Methodology")
    title_cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    title_cell.fill      = PatternFill("solid", fgColor="1F3864")
    title_cell.alignment = CENTER
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)

    # ── Row 2: Section group headers ───────────────────────────────────────
    ws.row_dimensions[2].height = 18
    sections = [
        ("Structure Analysis",  1,  6,  "2E75B6"),
        ("Failure Analysis",    7,  10, "C55A11"),
        ("Risk Assessment",     11, 17, "375623"),
    ]
    for label, c_start, c_end, color in sections:
        cell = ws.cell(row=2, column=c_start, value=label)
        cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=9)
        cell.fill      = PatternFill("solid", fgColor=color)
        cell.alignment = CENTER
        if c_start != c_end:
            ws.merge_cells(start_row=2, start_column=c_start, end_row=2, end_column=c_end)

    # ── Row 3: Column headers ──────────────────────────────────────────────
    ws.row_dimensions[3].height = 36
    for col_idx, (header, width, section, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font      = DARK_BOLD
        cell.fill      = SECTION_FILLS[section]
        cell.alignment = CENTER
        cell.border    = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Freeze panes below header rows ─────────────────────────────────────
    ws.freeze_panes = "A4"

    # ── Data rows ──────────────────────────────────────────────────────────
    for row_num, row in enumerate(rows, start=4):
        ws.row_dimensions[row_num].height = 48

        rpn = row.rpn or 0
        row_bg = (
            PatternFill("solid", fgColor="FFD7D7") if rpn >= 200 else   # red tint
            PatternFill("solid", fgColor="FFF2CC") if rpn >= 100 else   # amber tint
            None
        )

        values = [
            row.focus_element,
            row.focus_function,
            row.lower_element,
            row.lower_function,
            row.higher_element,
            row.higher_function,
            row.failure_mode,
            row.noise_factor,
            row.failure_cause,
            row.failure_effect,
            row.prevention_methods,
            row.detection_methods,
            row.severity,
            row.occurrence,
            row.detection,
            row.rpn,
            row.action_priority,
        ]

        for col_idx, (value, (_, _, section, _)) in enumerate(zip(values, COLUMNS), start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.font   = NORMAL_FONT
            cell.border = THIN_BORDER

            is_numeric_col = col_idx >= 13  # S, O, D, RPN, AP
            cell.alignment = CENTER if is_numeric_col else LEFT

            # Background: row tint takes priority, else section alternating
            if row_bg:
                cell.fill = row_bg
            else:
                # Subtle alternating section tint on even rows
                if row_num % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor="F5F8FF") if section == "structure" \
                           else PatternFill("solid", fgColor="FFF8F5") if section == "failure" \
                           else PatternFill("solid", fgColor="F5FFF8")

            # RPN cell: bold + colour
            if col_idx == 16 and value is not None:
                cell.font = Font(name="Arial", bold=True, size=9,
                                 color="CC0000" if rpn >= 200 else
                                       "B85C00" if rpn >= 100 else "1F5C1F")

            # AP cell: bold
            if col_idx == 17 and value:
                cell.font = Font(name="Arial", bold=True, size=9,
                                 color="CC0000" if value == "H" else
                                       "B85C00" if value == "M" else "1F5C1F")

    # ── Auto-filter on header row ──────────────────────────────────────────
    ws.auto_filter.ref = f"A3:{get_column_letter(num_cols)}{len(rows) + 3}"

    return wb


# ── Endpoint ──────────────────────────────────────────────────────────────────

@router.post("/export/template")
def export_template(req: ExportRequest):
    """
    Generate a styled DFMEA worksheet as .xlsx and return as a
    streaming file download.
    """
    wb = _build_workbook(req.rows)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="DFMEA_filled.xlsx"'},
    )
